"""Tests for the synthetic document fixture generator.

Covers docs/superpowers/test-plans/02-test-corpus-tests.md §1.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

import openpyxl
import pdfplumber
import pypdf
import pytest

from scripts.make_fixtures import (
    ANNUAL_LEAVE_DAYS,
    BUDGET_SHEET_NAMES,
    SALARY_BANDS,
    build_all,
)

EXPECTED_FILENAMES = {
    "handbook.pdf",
    "salary_bands.pdf",
    "ragged_salary_grid.pdf",
    "nda.docx",
    "expense_policy.docx",
    "wrapped_table.docx",
    "budget.xlsx",
    "corrupt.pdf",
    "scanned.pdf",
    "duplicate.pdf",
}


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _extract_ragged_row_lengths(pdf_path: Path) -> list[int]:
    """Group words on the page into text lines by y-position, then split
    each line into 'columns' on gaps wider than a small threshold. This is
    a deliberately naive column detector — it only needs to prove the grid
    is genuinely irregular, not to be a production table parser.
    """
    with pdfplumber.open(pdf_path) as pdf:
        words = pdf.pages[0].extract_words()

    lines: dict[float, list[dict]] = {}
    for w in words:
        key = round(w["top"], 0)
        lines.setdefault(key, []).append(w)

    row_lengths = []
    for _, line_words in sorted(lines.items()):
        line_words.sort(key=lambda w: w["x0"])
        columns = 1
        for prev, cur in zip(line_words, line_words[1:]):
            if cur["x0"] - prev["x1"] > 20:
                columns += 1
        row_lengths.append(columns)
    return row_lengths


class TestFixtureGenerator:
    def test_all_fixtures_written(self, tmp_path):
        paths = build_all(tmp_path)
        names = {p.name for p in paths}

        assert names == EXPECTED_FILENAMES
        for path in paths:
            assert path.exists()
            assert path.stat().st_size > 0

    def test_byte_reproducibility(self, tmp_path):
        out1 = tmp_path / "build1"
        out2 = tmp_path / "build2"
        build_all(out1)
        build_all(out2)

        for name in EXPECTED_FILENAMES:
            h1 = _sha256(out1 / name)
            h2 = _sha256(out2 / name)
            assert h1 == h2, f"{name} is not byte-reproducible across builds"

    def test_duplicate_matches_source(self, tmp_path):
        build_all(tmp_path)

        assert _sha256(tmp_path / "duplicate.pdf") == _sha256(
            tmp_path / "salary_bands.pdf"
        )

    def test_handbook_contains_known_leave_figure(self, tmp_path):
        build_all(tmp_path)

        reader = pypdf.PdfReader(tmp_path / "handbook.pdf")
        text = "\n".join(page.extract_text() for page in reader.pages)

        assert str(ANNUAL_LEAVE_DAYS) in text

    def test_salary_table_has_every_band(self, tmp_path):
        build_all(tmp_path)

        reader = pypdf.PdfReader(tmp_path / "salary_bands.pdf")
        text = "\n".join(page.extract_text() for page in reader.pages)

        for band, _min_val, mid_val, _max_val in SALARY_BANDS:
            assert band in text
            assert str(mid_val) in text

    def test_ragged_grid_is_genuinely_ragged(self, tmp_path):
        build_all(tmp_path)

        row_lengths = _extract_ragged_row_lengths(tmp_path / "ragged_salary_grid.pdf")

        assert len(row_lengths) > 1
        assert len(set(row_lengths)) > 1, (
            f"expected differing column counts across rows, got {row_lengths}"
        )

    def test_scanned_pdf_yields_no_text(self, tmp_path):
        build_all(tmp_path)

        reader = pypdf.PdfReader(tmp_path / "scanned.pdf")
        text = "\n".join(page.extract_text() for page in reader.pages)

        assert text.strip() == ""

    def test_corrupt_pdf_fails_to_parse(self, tmp_path):
        build_all(tmp_path)

        with pytest.raises(Exception):
            pypdf.PdfReader(tmp_path / "corrupt.pdf")

    def test_budget_workbook_sheets(self, tmp_path):
        build_all(tmp_path)

        wb = openpyxl.load_workbook(tmp_path / "budget.xlsx", data_only=True)

        assert set(BUDGET_SHEET_NAMES).issubset(set(wb.sheetnames))
        numeric_found = False
        for name in BUDGET_SHEET_NAMES:
            ws = wb[name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)):
                        numeric_found = True
        assert numeric_found
